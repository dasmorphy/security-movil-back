

import os
from typing import Counter
from openpyxl import load_workbook
from datetime import datetime
from swagger_server.models.db.logbook_entry import LogbookEntry
from swagger_server.models.db.logbook_out import LogbookOut
from swagger_server.models.request_post_logbook_entry import RequestPostLogbookEntry
from swagger_server.models.request_post_logbook_out import RequestPostLogbookOut
from swagger_server.repository.logbook_repository import LogbookRepository
from openpyxl.styles import Font, PatternFill

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Spacer, Image
from collections import OrderedDict, defaultdict


class LogbookUseCase:

    def __init__(self, logbook_repository: LogbookRepository):
        self.logbook_repository = logbook_repository

    def post_logbook_entry(self, body, images, internal, external) -> None:
        logbook_entry = LogbookEntry(
            unity_id=body['id_unity'],
            category_id=body['id_category'],
            group_business_id=body['id_group_business'],
            shipping_guide=body['shipping_guide'],
            description=body['description'],
            quantity=body['quantity'],
            weight=body['weight'],
            provider=body['provider'],
            truck_license=body['truck_license'],
            name_driver=body['name_driver'],
            destiny_intern=body['destiny_intern'],
            authorized_by=body['authorized_by'],
            observations=body['observations'],
            created_by=body['created_by'],
            updated_by=body['created_by'],
            name_user=body['name_user'],
            workday=body['workday']
        )

        self.logbook_repository.post_logbook_entry(logbook_entry, images, internal, external)

    def post_logbook_out(self, body, images, internal, external) -> None:
        logbook_out = LogbookOut(
            unity_id=body['id_unity'],
            category_id=body['id_category'],
            group_business_id=body['id_group_business'],
            shipping_guide=body['shipping_guide'],
            quantity=body['quantity'],
            weight=body['weight'],
            truck_license=body['truck_license'],
            name_driver=body['name_driver'],
            person_withdraws=body['person_withdraws'],
            destiny=body['destiny'],
            authorized_by=body['authorized_by'],
            observations=body['observations'],
            created_by=body['created_by'],
            updated_by=body['created_by'],
            name_user=body['name_user'],
            workday=body['workday']
        )

        self.logbook_repository.post_logbook_out(logbook_out, images, internal, external)

    def get_all_categories(self, internal, external):
        return self.logbook_repository.get_all_categories(internal, external)
    
    def get_all_unities(self, internal, external):
        return self.logbook_repository.get_all_unities(internal, external)
    
    def get_all_sector(self, internal, external):
        return self.logbook_repository.get_all_sectores(internal, external)
    
    def get_sector_by_business(self, id_business, internal, external):
        return self.logbook_repository.get_sector_by_business(id_business, internal, external)

    def get_user_info(datos):
        return 

    def get_sector_by_id(self, id_sector, internal, external):
        return self.logbook_repository.get_sector_by_id(id_sector, internal, external)
    
    def get_group_business_by_id_business(self, id_business, internal, external):
        return self.logbook_repository.get_group_business_by_id_business(id_business, internal, external)

    def get_logbooks_entry(self, headers, params, internal, external):
        groups = headers.get("groups_business_id")
        sectors = headers.get("sectors")
        workday = headers.get("workday")
        filters = {
            "user": headers.get("user"),
            "groups_business_id": [int(x) for x in groups.split(",")] if groups else [],
            "start_date": params.get("start_date"),
            "end_date": params.get("end_date"),
            "id_business": params.get("id_business"),
            "sector_id": [int(x) for x in sectors.split(",")] if sectors else [],
            "workday": [(x) for x in workday.split(",")] if workday else [],
        }
        rows = self.logbook_repository.get_all_logbook_entry(filters, internal, external)

        results = [
            {
                "id_logbook_entry": c.id_logbook_entry,
                "unity_id": c.unity_id,
                "group_name": group_name,
                "name_user": c.name_user,
                "category_id": c.category_id,
                "group_business_id": c.group_business_id,
                "shipping_guide": c.shipping_guide,
                "description": c.description,
                "quantity": c.quantity,
                "weight": c.weight,
                "provider": c.provider,
                "destiny_intern": c.destiny_intern,
                "authorized_by": c.authorized_by,
                "truck_license": c.truck_license,
                "name_driver": c.name_driver,
                "observations": c.observations,
                "created_at": c.created_at,
                "updated_at": c.updated_at,
                "created_by": c.created_by,
                "updated_by": c.updated_by,
                "workdday": c.workday,
                "id_sector": id_sector,
                "name_sector": name_sector,
                "images": images or []
            }
            for c, group_name, id_sector, name_sector, images in rows
        ]

        return results
    
    def get_logbooks_out(self, headers, params, internal, external):
        groups = headers.get("groups_business_id")
        sectors = headers.get("sectors")
        workday = headers.get("workday")
        filters = {
            "user": headers.get("user"),
            "groups_business_id": [int(x) for x in groups.split(",")] if groups else [],
            "start_date": params.get("start_date"),
            "end_date": params.get("end_date"),
            "sector_id": [int(x) for x in sectors.split(",")] if sectors else [],
            "workday": [(x) for x in workday.split(",")] if workday else [],
            "id_business": params.get("id_business"),
        }
        rows = self.logbook_repository.get_all_logbook_out(filters, internal, external)

        results = [
            {
                "id_logbook_out": c.id_logbook_out,
                "unity_id": c.unity_id,
                "category_id": c.category_id,
                "name_user": c.name_user,
                "group_name": group_name,
                "group_business_id": c.group_business_id,
                "shipping_guide": c.shipping_guide,
                "quantity": c.quantity,
                "weight": c.weight,
                "truck_license": c.truck_license,
                "name_driver": c.name_driver,
                "person_withdraws": c.person_withdraws,
                "destiny": c.destiny,
                "authorized_by": c.authorized_by,
                "observations": c.observations,
                "created_at": c.created_at,
                "updated_at": c.updated_at,
                "created_by": c.created_by,
                "updated_by": c.updated_by,
                "workdday": c.workday,
                "id_sector": id_sector,
                "name_sector": name_sector,
                "images": images or []
            }
            for c, group_name, id_sector, name_sector, images in rows
        ]

        return results
    
    def get_history_logbooks(self, headers, params, internal, external):
        rows_entry = self.get_logbooks_entry(headers, params, internal, external)
        rows_out = self.get_logbooks_out(headers, params, internal, external)

        rows = rows_entry + rows_out

        rows.sort(
            key=lambda x: x["created_at"],
            reverse=True
        )

        return rows
    
    def get_resume_graphs(self, headers, params, internal, external):
        rows_entry = self.get_logbooks_entry(headers, params, internal, external)
        rows_out = self.get_logbooks_out(headers, params, internal, external)

        all_rows = rows_entry + rows_out

        if len(all_rows) == 0:
            return {
                "total_entrada": 0,
                "total_salida": 0,
                "categorias": []
            }

        # Totales generales
        total_entry = len(rows_entry)
        total_out = len(rows_out)

        categories = self.get_all_categories(internal, external)

        # Mapa {id: nombre}
        category_map = {
            c["id_category"]: c["name_category"]
            for c in categories
        }

        # Contadores por categoría
        entry_counter = Counter(row["category_id"] for row in rows_entry)
        out_counter = Counter(row["category_id"] for row in rows_out)

        # =========================
        # POR CANTIDAD (quantity)
        # =========================
        entry_quantity = defaultdict(int)
        out_quantity = defaultdict(int)

        for row in rows_entry:
            entry_quantity[row["category_id"]] += row.get("quantity", 0)

        for row in rows_out:
            out_quantity[row["category_id"]] += row.get("quantity", 0)

        # Todas las categorías presentes
        all_category_ids = (
            set(entry_counter.keys())
            | set(out_counter.keys())
            | set(entry_quantity.keys())
            | set(out_quantity.keys())
        )

        categorias = []
        categorias_cantidad = []


        for category_id in all_category_ids:
            categoria_nombre = category_map.get(category_id, "Sin categoría")

            # Por registros
            entrada_reg = entry_counter.get(category_id, 0)
            salida_reg = out_counter.get(category_id, 0)

            categorias.append({
                "categoria": categoria_nombre,
                "entrada": entrada_reg,
                "salida": salida_reg,
                "total": entrada_reg + salida_reg
            })

            # Por quantity
            entrada_qty = entry_quantity.get(category_id, 0)
            salida_qty = out_quantity.get(category_id, 0)

            categorias_cantidad.append({
                "categoria": categoria_nombre,
                "entrada": entrada_qty,
                "salida": salida_qty,
                "total": entrada_qty + salida_qty
            })

        return {
            "total_entrada": total_entry,
            "total_salida": total_out,
            "categorias": categorias,
            "categorias_cantidad": categorias_cantidad
        }
    
    def post_report_generated(self, datos, internal, external) -> None:
        self.logbook_repository.post_report_generated(datos, internal, external)


    def generate_excel(self, datos, output_path, internal, external):
        # '2026-01-27 00:00:00', '2026-01-28 00:00:00'
        now = datetime.now()
        date = now.strftime("%d/%m/%Y")
        time = now.strftime("%H:%M:%S")

        logbook_entry_rows = self.logbook_repository.get_logbook_resume(internal, external, LogbookEntry)
        logbook_out_rows = self.logbook_repository.get_logbook_resume(internal, external, LogbookOut)
        # business_user, name_user = self.get_user_info(datos)

        print(logbook_entry_rows)
        print(logbook_out_rows)

        resultado = {}

        for row in logbook_out_rows:
            grupo = row[1]          # GroupBusiness.name
            cat_id = row[2]         # Category.id_category

            if grupo not in resultado:
                resultado[grupo] = {
                    "categorias": {}
                }

            resultado[grupo]["categorias"][cat_id] = {
                "categoria": row[3],
                "unidad": row[5],
                "salida": row[4],
                "entrada": 0
            }


        for row in logbook_entry_rows:
            grupo = row[1]
            cat_id = row[2]

            if grupo not in resultado:
                resultado[grupo] = {
                    "categorias": {}
                }

            if cat_id not in resultado[grupo]["categorias"]:
                resultado[grupo]["categorias"][cat_id] = {
                    "categoria": row[3],
                    "unidad": row[5],
                    "salida": 0,
                    "entrada": row[4]
                }
            else:
                resultado[grupo]["categorias"][cat_id]["entrada"] = row[4]


        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(
            BASE_DIR,
            "templates",
            "template_report.xlsx"
        )

        wb = load_workbook(template_path)
        ws = wb.active

        # ws_copia = wb.copy_worksheet(wb)
        # ws_copia.title = "REPORTE 2"

        # Cabecera
        ws["C7"] = date
        ws["C8"] = datos["localidad"]
        ws["C9"] = datos["puesto_control"]
        ws["C10"] = datos["agente"]
        ws["B14"] = "TAURA (TOTAL)"

        fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"
        )

        ws["B14"].fill = fill
        ws["B14"].font = Font(bold=True)

        ws["F7"] = datos["ref"]
        ws["F8"] = time

        fila_inicio = 15

        for grupo, data in resultado.items():
            ws[f"B{fila_inicio}"] = grupo.upper()
            ws[f"B{fila_inicio}"].font = Font(bold=True)

            fila_inicio += 1

            for item in data["categorias"].values():
                ws[f"B{fila_inicio}"] = f"TOTAL {item['categoria'].upper()}"
                ws[f"C{fila_inicio}"] = item["salida"]
                ws[f"D{fila_inicio}"] = item["unidad"]
                ws[f"E{fila_inicio}"] = item["entrada"]
                ws[f"F{fila_inicio}"] = item["unidad"]
                
                fila_inicio += 1
            
            fila_inicio += 1

        wb.save(output_path)


    def agrupar_por_categoria(self, logbook_entry_rows):
        agrupado = OrderedDict()

        for row in logbook_entry_rows:
            category_id = row[2]

            if category_id not in agrupado:
                # Guardamos la fila base
                agrupado[category_id] = list(row)
            else:
                # Sumamos SOLO la cantidad (índice 4)
                agrupado[category_id][4] += row[4]

        # Convertimos de vuelta a tuplas
        return [tuple(row) for row in agrupado.values()]


    def generate_pdf(self, filters, output_path, internal, external):
        now = datetime.now()
        date = now.strftime("%d/%m/%Y")
        time = now.strftime("%H:%M:%S")
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(BASE_DIR, "templates", "logo_report.png")
        sectors = filters[0].get('id_sector')
        groups_business_id = filters[0].get('groups_business_id')
        
        logo = Image(
            logo_path,
            width=230,
            height=60
        )

        filters_dict = {
            "sector_id": [int(x) for x in sectors.split(",")] if sectors else [],
            "groups_business_id": [int(x) for x in groups_business_id.split(",")] if groups_business_id else [],
            "fecha_inicio": filters[1].get('start_date'),
            "fecha_fin": filters[1].get('end_date')
        }

        logbook_entry_rows_original = self.logbook_repository.get_logbook_resume(internal, external, LogbookEntry, filters_dict)
        logbook_out_rows_original = self.logbook_repository.get_logbook_resume(internal, external, LogbookOut, filters_dict)
        
        logbook_entry_rows = self.agrupar_por_categoria(logbook_entry_rows_original)
        logbook_out_rows = self.agrupar_por_categoria(logbook_out_rows_original)

        print(logbook_entry_rows)
        print(logbook_out_rows)

        # Determinar si se agrupa por sector o por grupos
        if not sectors:
            # Agrupar por localidad (sector)
            self._generate_pdf_by_locality(
                logo, date, time, 
                logbook_entry_rows, logbook_out_rows, 
                output_path
            )
        else:
            # Agrupar por grupos de negocio
            self._generate_pdf_by_groups(
                logo, date, time, 
                logbook_entry_rows, logbook_out_rows, 
                sectors, output_path, internal, external
            )


    def _generate_pdf_by_locality(self, logo, date, time, logbook_entry_rows, logbook_out_rows, output_path):
        """
        Genera el PDF agrupando por localidad (sector)
        """
        resultado_por_localidad = {}

        # Procesar salidas
        for row in logbook_out_rows:
            # row = (id, grupo, cat_id, categoria, cantidad, unidad, sector_id, localidad)
            localidad = row[7]
            cat_id = row[2]

            if localidad not in resultado_por_localidad:
                resultado_por_localidad[localidad] = {
                    "categorias": {}
                }

            if cat_id not in resultado_por_localidad[localidad]["categorias"]:
                resultado_por_localidad[localidad]["categorias"][cat_id] = {
                    "categoria": row[3],
                    "unidad": row[5],
                    "salida": 0,
                    "entrada": 0
                }
            
            resultado_por_localidad[localidad]["categorias"][cat_id]["salida"] += row[4]

        # Procesar entradas
        for row in logbook_entry_rows:
            localidad = row[7]
            cat_id = row[2]
            
            if localidad not in resultado_por_localidad:
                resultado_por_localidad[localidad] = {
                    "categorias": {}
                }

            if cat_id not in resultado_por_localidad[localidad]["categorias"]:
                resultado_por_localidad[localidad]["categorias"][cat_id] = {
                    "categoria": row[3],
                    "unidad": row[5],
                    "salida": 0,
                    "entrada": 0
                }
            
            resultado_por_localidad[localidad]["categorias"][cat_id]["entrada"] += row[4]

        # Crear el PDF
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()
        elements = []

        # Encabezado
        elements.append(logo)
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Fecha: {date}", styles["Normal"]))
        elements.append(Paragraph(f"Hora: {time}", styles["Normal"]))
        elements.append(Paragraph("Localidad: TODAS", styles["Normal"]))
        elements.append(Paragraph("Puesto de control: GARITA DE SEGURIDAD", styles["Normal"]))
        elements.append(Paragraph("Agente: ", styles["Normal"]))
        elements.append(Paragraph("REP#: RP2026-010", styles["Normal"]))
        elements.append(Spacer(1, 12))
        elements.append(Spacer(1, 12))

        # Tabla
        table_data = [
            ["Descripción", "Salida", "Unidad", "Entrada", "Unidad"]
        ]

        localidad_row_indexes = []

        # Agregar cada localidad
        for localidad, data in sorted(resultado_por_localidad.items()):
            localidad_row_index = len(table_data)
            localidad_row_indexes.append(localidad_row_index)

            # Fila con la localidad
            table_data.append([
                f"{localidad.upper()} (TOTAL)", "", "", "", ""
            ])

            # Agregar categorías de esta localidad
            for cat_id in sorted(data["categorias"].keys()):
                item = data["categorias"][cat_id]
                table_data.append([
                    f"TOTAL {item['categoria'].upper()}",
                    item["salida"],
                    item["unidad"],
                    item["entrada"],
                    item["unidad"]
                ])

            # Espacio después de cada localidad
            table_data.append(["", "", "", "", ""])

        table = Table(table_data, colWidths=[180, 60, 60, 60, 60])

        style_commands = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ]

        # Estilo para cada localidad (amarillo)
        for row in localidad_row_indexes:
            style_commands.extend([
                ("BACKGROUND", (0, row), (-1, row), colors.yellow),
                ("FONTNAME", (0, row), (0, row), "Helvetica-Bold"),
            ])

        table.setStyle(TableStyle(style_commands))

        elements.append(table)
        doc.build(elements)


    def _generate_pdf_by_groups(self, logo, date, time, logbook_entry_rows, logbook_out_rows, sectors, output_path, internal, external):
        """
        Genera el PDF agrupando por grupos de negocio
        """
        localidad = "TODAS"

        if sectors:
            sector_found = self.logbook_repository.get_sector_by_id(sectors[0], internal, external)
            localidad = sector_found.get('name')

        resultado = {}

        # Procesar salidas
        for row in logbook_out_rows:
            grupo = row[1]
            cat_id = row[2]

            if grupo not in resultado:
                resultado[grupo] = {
                    "categorias": {}
                }

            resultado[grupo]["categorias"][cat_id] = {
                "categoria": row[3],
                "unidad": row[5],
                "salida": row[4],
                "entrada": 0
            }

        # Procesar entradas
        for row in logbook_entry_rows:
            grupo = row[1]
            cat_id = row[2]
            
            if grupo not in resultado:
                resultado[grupo] = {
                    "categorias": {}
                }

            if cat_id not in resultado[grupo]["categorias"]:
                resultado[grupo]["categorias"][cat_id] = {
                    "categoria": row[3],
                    "unidad": row[5],
                    "salida": 0,
                    "entrada": row[4]
                }
            else:
                resultado[grupo]["categorias"][cat_id]["entrada"] = row[4]

        # Calcular totales generales por categoría
        totales_generales = {}
        
        for grupo, data in resultado.items():
            for cat_id, item in data["categorias"].items():
                if cat_id not in totales_generales:
                    totales_generales[cat_id] = {
                        "categoria": item["categoria"],
                        "unidad": item["unidad"],
                        "salida": 0,
                        "entrada": 0
                    }
                
                totales_generales[cat_id]["salida"] += item["salida"]
                totales_generales[cat_id]["entrada"] += item["entrada"]

        # Crear el PDF
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()
        elements = []

        # Encabezado
        elements.append(logo)
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"Fecha: {date}", styles["Normal"]))
        elements.append(Paragraph(f"Hora: {time}", styles["Normal"]))
        elements.append(Paragraph(f"Localidad: {localidad}", styles["Normal"]))
        elements.append(Paragraph("Puesto de control: GARITA DE SEGURIDAD", styles["Normal"]))
        elements.append(Paragraph("Agente: ", styles["Normal"]))
        elements.append(Paragraph("REP#: RP2026-010", styles["Normal"]))
        elements.append(Spacer(1, 12))
        elements.append(Spacer(1, 12))

        # Tabla
        table_data = [
            ["Descripción", "Salida", "Unidad", "Entrada", "Unidad"]
        ]

        group_row_indexes = []
        localidad_row_indexes = []

        # LOCALIDAD con totales
        localidad_row_index = len(table_data)
        localidad_row_indexes.append(localidad_row_index)

        table_data.append([
            f"{localidad.upper()} (TOTAL)", "", "", "", ""
        ])

        # Agregar totales generales
        for cat_id in sorted(totales_generales.keys()):
            item = totales_generales[cat_id]
            table_data.append([
                f"TOTAL {item['categoria'].upper()}",
                item["salida"],
                item["unidad"],
                item["entrada"],
                item["unidad"]
            ])

        table_data.append(["", "", "", "", ""])

        # Agregar grupos individuales
        for grupo, data in resultado.items():
            group_row_index = len(table_data)
            group_row_indexes.append(group_row_index)

            table_data.append([
                grupo.upper(), "", "", "", ""
            ])

            for item in data["categorias"].values():
                table_data.append([
                    f"TOTAL {item['categoria'].upper()}",
                    item["salida"],
                    item["unidad"],
                    item["entrada"],
                    item["unidad"]
                ])

            table_data.append(["", "", "", "", ""])

        table = Table(table_data, colWidths=[180, 60, 60, 60, 60])

        style_commands = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ]

        # Estilo LOCALIDAD (amarillo)
        for row in localidad_row_indexes:
            style_commands.extend([
                ("BACKGROUND", (0, row), (-1, row), colors.yellow),
                ("FONTNAME", (0, row), (0, row), "Helvetica-Bold"),
            ])

        # Estilo GRUPO (solo negrita)
        for row in group_row_indexes:
            style_commands.append(
                ("FONTNAME", (0, row), (0, row), "Helvetica-Bold")
            )

        table.setStyle(TableStyle(style_commands))

        elements.append(table)
        doc.build(elements)
